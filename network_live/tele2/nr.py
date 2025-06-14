from datetime import date

import openpyxl
from network_live.ftp import download_ftp_logs, get_date
from network_live.physical_params import add_physical_params
from point_in_region import find_region_by_coordinates

nl_headears = {
    'NE Name': 'site_name',
    'Cell Name': 'cell_name',
    'CI': 'cellLocalId',
    'gNodeB ID': 'gNBId',
    'eCI': 'nCI',
    'Physical Cell ID': 'nRPCI',
    'TAC': 'nRTAC',
    'Logical Root Sequence Index': 'rachRootSequence',
    'Minimum RX Level(2dBm)': 'qRxLevMin',
    'DL NARFCN': 'arfcnDL',
    'Frequency Band': 'bSChannelBwDL',
    'Max Transmit Power(0.1dBm)': 'configuredMaxTxPower',
}


def get_header(sheet, cell):
    header = sheet.cell(row=1, column=cell.column)
    return header.value


def parse_nr(log_path, atoll_data):
    nr_cells = []
    headers = nl_headears.keys()
    work_book = openpyxl.load_workbook(log_path)
    sheet = work_book.active


    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
        nr_cell = {
            'subnetwork': 'Tele2',
            'cellState': None,
            'ip_address': None,
            'vendor': 'Huawei',
            'oss': 'Tele2',
            'ssbFrequency': None,
            'insert_date': date.today(),
            'sharing_type': None,
        }
        for cell in row:
            header = get_header(sheet, cell)
            if header in headers:
                nr_cell[nl_headears[header]] = cell.value
            if cell.column == sheet.max_column:
                cell_with_phys_params = add_physical_params(atoll_data, 'NR', nr_cell)
                try:
                    cell_with_phys_params['region'] = find_region_by_coordinates(
                        (cell_with_phys_params['longitude'], cell_with_phys_params['latitude']),
                    )
                except TypeError:
                    cell_with_phys_params['region'] = None

                nr_cells.append(cell_with_phys_params)

    work_book.close()

    return nr_cells


def nr_main(atoll_data):
    download_ftp_logs('tele2_nr', is_unzip=False)

    date = get_date('tele2')
    log_path = f'logs/tele2/5G_Kcell_CM-{date}.xlsx'
    cells = parse_nr(log_path, atoll_data)
    return cells


if __name__ == '__main__':
    nr_main()
